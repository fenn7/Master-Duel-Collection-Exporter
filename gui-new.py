import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys
import subprocess
from typing import List, Tuple
import threading
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import requests
try:
    import win32com.client as win32
    HAS_WIN32 = True
except Exception:
    HAS_WIN32 = False

print("Launching Master Duel Collection Exporter!")

ARROW_MAP = {
    "Top": "↑","Bottom": "↓","Left": "←","Right": "→",
    "Top-Left": "↖","Top-Right": "↗","Bottom-Left": "↙","Bottom-Right": "↘"
}
ANSI_TO_TAG = {
    '\x1b[0m': None,'\x1b[94m': 'blue','\x1b[97m': 'white','\x1b[92m': 'green',
    '\x1b[93m': 'yellow','\x1b[91m': 'red','\x1b[96m': 'cyan','\x1b[1;97m': 'bold_white',
    '\x1b[1;96m': 'bold_cyan','\x1b[1;93m': 'bold_yellow','\x1b[1;94m': 'bold_blue','\x1b[4;97m': 'underline_white'
}
ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')

def get_card_info(canonical_name: str) -> dict:
    """Fetch card info from YGOPRODECK API."""
    if not canonical_name:
        return {}
    url = f"https://db.ygoprodeck.com/api/v7/cardinfo.php?name={quote(canonical_name)}&misc=yes"
    try:
        r = requests.get(url, timeout=5.0)
        r.raise_for_status()
        data = r.json()
        if "data" in data and data["data"]:
            return data["data"][0]
        return {"error": data.get("error", "not found")}
    except Exception:
        return {}

def get_card_rarity_from_info(card_info: dict) -> str:
    """Extract Master Duel rarity from card info misc_info."""
    for misc in card_info.get("misc_info", []):
        if "md_rarity" in misc:
            return {"Common": "N ","Rare": "R ","Super Rare": "SR","Ultra Rare": "UR"}.get(misc["md_rarity"], "N ")
    return "N "

def parse_ansi(text: str) -> tuple[str, List[Tuple[int, int, str]]]:
    """Strip ANSI codes and return clean text plus tag ranges suitable for Tk Text widget."""
    clean_parts = []
    tag_ranges = []
    pos = 0
    current_tag = None
    last_end = 0
    for m in ANSI_RE.finditer(text):
        s, e = m.span()
        if s > last_end:
            seg = text[last_end:s]
            clean_parts.append(seg)
            if current_tag:
                tag_ranges.append((pos, pos + len(seg), current_tag))
            pos += len(seg)
        code = m.group()
        current_tag = ANSI_TO_TAG.get(code, current_tag)
        if code == '\x1b[0m':
            current_tag = None
        last_end = e
    if last_end < len(text):
        seg = text[last_end:]
        clean_parts.append(seg)
        if current_tag:
            tag_ranges.append((pos, pos + len(seg), current_tag))
    return ''.join(clean_parts), tag_ranges

class MasterDuelExporterApp:
    """Tkinter GUI for exporting and viewing Master Duel collections."""
    def __init__(self, root):
        self.root = root
        self.root.title("Master Duel Collection Exporter")
        sw = self.root.winfo_screenwidth()
        self.root.geometry(f"500x750+{sw-500}+0")
        self.root.minsize(500, 750)
        self.scan_in_progress = False
        self.debug_mode = tk.BooleanVar(value=False)
        self.print_summary = tk.BooleanVar(value=False)
        self.game_resolution = tk.StringVar(value='1600x900')
        self.process = None
        self.execution_terminal = None
        self.load_terminal = None
        self.terminal_contents = {}
        try:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'favicon.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass
        self.style = ttk.Style()
        self.style.configure('TButton', font=('Arial', 12), padding=10)
        self.style.configure('Small.TButton', font=('Arial', 9), padding=5)
        self.style.configure('Title.TLabel', font=('Arial', 18, 'bold'))
        self.style.configure('Subtitle.TLabel', font=('Arial', 12))
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(expand=True, fill='both', padx=20, pady=20)
        self.show_main_menu()

    def clear_frame(self):
        """Remove all widgets from main frame."""
        for w in self.main_frame.winfo_children():
            w.destroy()

    def show_main_menu(self):
        """Render the main menu UI."""
        self.clear_frame()
        title = ttk.Label(self.main_frame, text="Master Duel Collection Exporter", style='Title.TLabel')
        title.pack(pady=(20, 40))
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(expand=True)
        new_btn = ttk.Button(button_frame, text="Create New Collection CSV", command=self.create_new_collection, width=25)
        new_btn.pack(pady=10)
        load_btn = ttk.Button(button_frame, text="Load a Collection CSV", command=self.load_existing_collection, width=25)
        load_btn.pack(pady=10)
        help_text = ttk.Label(self.main_frame, text="Contact Skybullet07 on Discord for help, or to report issues.", style='Subtitle.TLabel', foreground='gray50')
        help_text.pack(side=tk.BOTTOM, pady=10)

    def create_new_collection(self):
        """Show Create New Collection screen and controls."""
        self.clear_frame()
        self.default_save_dir = os.path.join(os.getcwd(), 'collection_csv')
        os.makedirs(self.default_save_dir, exist_ok=True)
        self.save_path = self.default_save_dir
        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill='x', pady=(0, 14))
        title = ttk.Label(header_frame, text="Create New Collection", style='Title.TLabel', anchor='center')
        title.pack(expand=True, fill='x', pady=(0, 2))
        back_btn = ttk.Button(header_frame, text="← Back", command=self._back_from_create, style='Small.TButton')
        back_btn.place(x=0, y=0, anchor='nw')
        info = ttk.Label(self.main_frame, text="Save your Master Duel collection to a new CSV file.", style='Subtitle.TLabel')
        info.pack(pady=(0, 7))
        save_selector_frame = ttk.LabelFrame(self.main_frame, text="Select save folder:", padding=4)
        save_selector_frame.pack(fill='x', padx=35, pady=7)
        entry_frame = ttk.Frame(save_selector_frame)
        entry_frame.pack(fill='x', padx=5, pady=5)
        self.save_entry = ttk.Entry(entry_frame)
        self.save_entry.insert(0, self.save_path)
        self.save_entry.pack(side='left', fill='x', expand=True, padx=(0, 5))
        browse_btn = ttk.Button(entry_frame, text="...", command=self.browse_save_location, style='Small.TButton', width=3)
        browse_btn.pack(side='right')
        instructions_frame = ttk.LabelFrame(self.main_frame, text="INSTRUCTIONS", padding=4)
        instructions_frame.pack(fill='x', padx=14, pady=(14, 4))
        instructions_text = (
            "1. Open Master Duel and click on the \"Deck\" button. \n"
            "2. Click on any of your Decks, then click \"Edit Deck\". \n"
            "3. Set the resolution to match your setup, then click \"Start Collection Scan\". \n"
            "4. DON'T close/minimise Master Duel or move your cursor during scan.\n"
            "5. Results will be saved to a CSV file in the selected folder."
        )
        ttk.Label(instructions_frame, text=instructions_text, justify='left').pack(anchor='w')
        resolution_frame = ttk.Frame(self.main_frame)
        resolution_frame.pack(anchor='center', padx=14, pady=(7, 4))
        ttk.Label(resolution_frame, text="Game Resolution:").grid(row=0, column=0, sticky='w', padx=(0,5), pady=2)
        game_combo = ttk.Combobox(resolution_frame, textvariable=self.game_resolution, values=['1280x720','1366x768','1440x810','1600x900','1920x1080'], state='readonly', width=12)
        game_combo.grid(row=0, column=1, sticky='w')
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(anchor='center', padx=35, pady=4)
        self.start_btn = ttk.Button(button_frame, text="Start Collection Scan", command=self.start_collection_scan, style='Small.TButton', padding=5)
        self.start_btn.pack(side='left', expand=False, padx=(0, 5))
        self.stop_btn = ttk.Button(button_frame, text="Stop Current Scan", command=self.stop_collection_scan, style='Small.TButton', padding=5, state='disabled')
        self.stop_btn.pack(side='left', expand=False, padx=(0, 5))
        options_frame = ttk.Frame(button_frame)
        options_frame.pack(side='left', padx=(10, 0))
        debug_check = ttk.Checkbutton(options_frame, text="Debug Mode", variable=self.debug_mode, command=self.log_debug_mode)
        debug_check.grid(row=0, column=0, sticky='w')
        summary_check = ttk.Checkbutton(options_frame, text="Print Summary", variable=self.print_summary, command=self.log_print_summary)
        summary_check.grid(row=1, column=0, sticky='w')
        terminal_frame = ttk.LabelFrame(self.main_frame, text="EXECUTION LOG", padding=4)
        terminal_frame.pack(fill='both', expand=True, padx=14, pady=(4, 14))
        self.execution_terminal = tk.Text(terminal_frame, bg='black', fg='white', font=('Consolas', 10), wrap=tk.WORD, state='disabled', padx=5, pady=5)
        for tag, cfg in {
            'timestamp': {'foreground':'#999999'}, 'info': {'foreground':'white'}, 'warning': {'foreground':'#ffff00'},
            'error': {'foreground':'#ff0000'}, 'debug': {'foreground':'#00ffff'}, 'blue': {'foreground':'#0080ff'},
            'white': {'foreground':'white'}, 'green': {'foreground':'#00ff00'}, 'yellow': {'foreground':'#ffff00'},
            'red': {'foreground':'#ff0000'}, 'cyan': {'foreground':'#00ffff'}
        }.items():
            self.execution_terminal.tag_configure(tag, **cfg)
        self.execution_terminal.tag_configure('bold_white', foreground='white', font=('Consolas', 10, 'bold'))
        self.execution_terminal.tag_configure('bold_cyan', foreground='#00ffff', font=('Consolas', 10, 'bold'))
        self.execution_terminal.tag_configure('bold_yellow', foreground='#ffff00', font=('Consolas', 10, 'bold'))
        self.execution_terminal.tag_configure('bold_blue', foreground='#0080ff', font=('Consolas', 10, 'bold'))
        self.execution_terminal.tag_configure('underline_white', foreground='white', underline=True)
        scrollbar = ttk.Scrollbar(terminal_frame, orient='vertical', command=self.execution_terminal.yview)
        self.execution_terminal.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.execution_terminal.pack(side='left', fill='both', expand=True)
        self.restore_terminal_content(self.execution_terminal, 'execution')
        if 'execution' not in self.terminal_contents:
            self.log("Terminal initialised.", "info")
            self.update_status("Ready to start a new export.")

    def load_existing_collection(self):
        """Show Load Collection screen and controls."""
        self.clear_frame()
        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill='x', pady=(0, 14))
        title = ttk.Label(header_frame, text="Load a Collection", style='Title.TLabel', anchor='center')
        title.pack(expand=True, fill='x', pady=(0, 2))
        back_btn = ttk.Button(header_frame, text="← Back", command=self._back_from_load, style='Small.TButton')
        back_btn.place(x=0, y=0, anchor='nw')
        info = ttk.Label(self.main_frame, text="Select an existing collection CSV file to view or edit:", style='Subtitle.TLabel')
        info.pack(pady=(0, 7))
        file_selector_frame = ttk.LabelFrame(self.main_frame, text="Select saved collection:", padding=4)
        file_selector_frame.pack(fill='x', padx=35, pady=7)
        file_frame = ttk.Frame(file_selector_frame)
        file_frame.pack(fill='x', padx=5, pady=5)
        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.pack(side='left', fill='x', expand=True, padx=(0, 5))
        browse_btn = ttk.Button(file_frame, text="...", command=self.browse_file, style='Small.TButton', width=3)
        browse_btn.pack(side='right')
        instructions_frame = ttk.LabelFrame(self.main_frame, text="INSTRUCTIONS", padding=4)
        instructions_frame.pack(fill='x', padx=14, pady=(14, 7))
        instructions_text = (
            "1. Select an existing collection CSV file using the browse button. \n"
            "2. Ensure that file is not already open! Click \"View Collection\" to view it. \n"
            "3. To update the data entries yourself, click on \"Update Collection\".\n"
            "4. Then follow the instructions to add, remove or amend collection data.\n"
            "5. NOTE: You will need to close and re-open the file to view your updates."
        )
        ttk.Label(instructions_frame, text=instructions_text, justify='left').pack(anchor='w')
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(fill='x', padx=35, pady=(14, 7))
        self.load_btn = ttk.Button(button_frame, text="View Collection", command=lambda: self.load_collection_file(self.file_entry.get()), style='Small.TButton', padding=5, state='disabled')
        self.load_btn.pack(side='left', expand=True, padx=(0, 5))
        self.update_btn = ttk.Button(button_frame, text="Update Collection", command=self.show_update_screen, style='Small.TButton', padding=5, state='disabled')
        self.update_btn.pack(side='right', expand=True, padx=(5, 0))
        terminal_frame = ttk.LabelFrame(self.main_frame, text="EXECUTION LOG", padding=4)
        terminal_frame.pack(fill='both', expand=True, padx=14, pady=(14, 14))
        self.load_terminal = tk.Text(terminal_frame, bg='black', fg='white', font=('Consolas', 10), wrap=tk.WORD, state='disabled', padx=5, pady=5)
        for tag, cfg in {'timestamp': {'foreground':'#999999'}, 'info': {'foreground':'white'}, 'warning': {'foreground':'#ffff00'}, 'error': {'foreground':'#ff0000'}}.items():
            self.load_terminal.tag_configure(tag, **cfg)
        scrollbar = ttk.Scrollbar(terminal_frame, orient='vertical', command=self.load_terminal.yview)
        self.load_terminal.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.load_terminal.pack(side='left', fill='both', expand=True)
        self.restore_terminal_content(self.load_terminal, 'load')
        if 'load' not in self.terminal_contents:
            self.load_log("Terminal initialized.", "info")
            self.load_log("Ready to load saved CSV.", "info")
        self.update_status("Select a collection file to load")

    def browse_file(self):
        """Open file dialog to select a CSV file."""
        from tkinter import filedialog
        filename = filedialog.askopenfilename(title="Select Collection File", filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")], initialdir=os.getcwd())
        if filename:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, filename)
            self.current_csv = filename
            self.load_log("File selected: " + os.path.basename(filename), "info")
            if filename.lower().endswith('.csv'):
                self.load_btn['state'] = 'normal'
                self.update_btn['state'] = 'normal'
            else:
                self.load_btn['state'] = 'disabled'
                self.update_btn['state'] = 'disabled'

    def browse_save_location(self):
        """Open directory dialog to choose where to save the CSV."""
        from tkinter import filedialog
        initial_dir = self.save_entry.get() or self.default_save_dir
        directory = filedialog.askdirectory(title="Select Save Folder", mustexist=False, initialdir=initial_dir)
        if directory:
            self.save_entry.delete(0, tk.END)
            self.save_entry.insert(0, directory)

    def load_collection_file(self, filepath):
        """Load an existing collection file and open a formatted XLSX copy."""
        if not filepath:
            self.load_log("Please select a file first", "error")
            return
        if not os.path.exists(filepath):
            self.load_log(f"File not found: {os.path.basename(filepath)}", "error")
            return
        self.load_log(f"Opening file: {os.path.basename(filepath)}", "info")
        self.current_csv = filepath
        self._open_file(filepath)
        if filepath.lower().endswith('.csv'):
            self.update_btn['state'] = 'normal'

    def _open_file(self, filepath):
        """Convert CSV to formatted XLSX and open it with OS default app."""
        filepath = os.path.abspath(filepath)
        try:
            base, ext = os.path.splitext(filepath)
            ext = (ext or "").lower()
            if ext in (".xlsx", ".xlsm", ".xls"):
                try:
                    os.startfile(filepath)
                    self.load_log("Opened existing Excel file", "info")
                    return
                except Exception:
                    try:
                        import subprocess, sys
                        if sys.platform.startswith("darwin"):
                            subprocess.call(["open", filepath])
                        else:
                            subprocess.call(["xdg-open", filepath])
                        self.load_log("Opened existing spreadsheet with default application", "info")
                        return
                    except Exception as e:
                        self.load_log(f"Failed to open existing spreadsheet: {e}", "error")
                        return
            formatted_dir = os.path.join(os.path.dirname(filepath), 'formatted')
            os.makedirs(formatted_dir, exist_ok=True)
            filename = os.path.basename(filepath)
            name = os.path.splitext(filename)[0]
            out_xlsx = os.path.join(formatted_dir, f"FORMATTED_{name}.xlsx")
            try:
                df = pd.read_csv(filepath, dtype=str, keep_default_na=False, na_values=[])
            except Exception:
                try:
                    df = pd.read_csv(filepath, dtype=str, encoding="latin-1", keep_default_na=False, na_values=[])
                except Exception as e:
                    self.load_log(f"Failed to read CSV: {e}", "error")
                    return
            try:
                df.to_excel(out_xlsx, index=False, engine="openpyxl")
            except Exception as e:
                self.load_log(f"Failed to write XLSX: {e}", "error")
                return
            try:
                wb = load_workbook(out_xlsx)
                ws = wb.active
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'E{row}']
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.isdigit():
                        cell.value = int(cell.value)
                MAX_WIDTH = 250
                PAD = 1
                for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
                    max_len = 0
                    for cell_val in col:
                        if cell_val is None:
                            continue
                        s = str(cell_val)
                        longest = max((len(line) for line in s.splitlines()), default=0)
                        if longest > max_len:
                            max_len = longest
                    raw_width = max(3, min(MAX_WIDTH, int(max_len + PAD)))
                    width = max(3, int(round(raw_width * 0.95)))
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width
                wb.save(out_xlsx)
                if HAS_WIN32:
                    try:
                        excel = win32.gencache.EnsureDispatch('Excel.Application')
                        wb_excel = excel.Workbooks.Open(out_xlsx)
                        ws_excel = wb_excel.Worksheets(1)
                        ws_excel.Cells.Select()
                        ws_excel.Columns.AutoFit()
                        wb_excel.Save()
                        wb_excel.Close()
                        excel.Quit()
                    except Exception:
                        pass
            except Exception as e:
                self.load_log(f"Auto-fit failed but XLSX was created: {e}", "warning")
            try:
                os.startfile(out_xlsx)
                self.load_log(f"Opened {os.path.basename(out_xlsx)} in default application", "info")
            except Exception:
                try:
                    import subprocess, sys
                    if sys.platform.startswith("darwin"):
                        subprocess.call(["open", out_xlsx])
                    else:
                        subprocess.call(["xdg-open", out_xlsx])
                    self.load_log(f"Opened {os.path.basename(out_xlsx)} with default application", "info")
                except Exception as e:
                    self.load_log(f"Failed to open generated XLSX: {e}", "error")
        except Exception as e:
            self.load_log(f"Unexpected error in _open_file: {e}", "error")

    def load_log(self, message: str, level: str = "info"):
        """Append a colored log line to the load terminal (or stdout if missing)."""
        if self.load_terminal is None or not self.load_terminal.winfo_exists():
            print(f"[{level.upper()}] {message}")
            return
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.load_terminal.configure(state='normal')
        self.load_terminal.insert('end', f"[{timestamp}] ", 'timestamp')
        self.load_terminal.insert('end', f"{message}\n", level)
        self.load_terminal.see('end')
        self.load_terminal.configure(state='disabled')
        self.root.update_idletasks()
        print(f"[{level.upper()}] [{timestamp}] {message}")

    def log(self, message: str, level: str = "info"):
        """Append a colored log line to the execution terminal (or stdout if missing)."""
        if self.execution_terminal is None or not self.execution_terminal.winfo_exists():
            print(f"[{level.upper()}] {message}")
            return
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.execution_terminal.configure(state='normal')
        self.execution_terminal.insert('end', f"[{timestamp}] ", 'timestamp')
        self.execution_terminal.insert('end', f"{message}\n", level)
        self.execution_terminal.see('end')
        self.execution_terminal.configure(state='disabled')
        self.root.update_idletasks()
        print(f"[{level.upper()}] [{timestamp}] {message}")

    def validate_number(self, value):
        """Return True if value is empty or an integer string."""
        return value == '' or value.isdigit()

    def parse_copies(self, copies_str):
        """Parse 'N [Basic A, Glossy B, Royal C]' into dict with totals."""
        if not copies_str or '[' not in copies_str:
            return {'total': 0, 'Basic': 0, 'Glossy': 0, 'Royal': 0}
        try:
            total_part, detail_part = copies_str.split(' [', 1)
            total = int(total_part.strip())
            detail = detail_part.rstrip(']')
            parts = [p.strip() for p in detail.split(',')]
            counts = {}
            for part in parts:
                finish, num = part.split()
                counts[finish] = int(num)
            return {'total': total, 'Basic': counts.get('Basic', 0), 'Glossy': counts.get('Glossy', 0), 'Royal': counts.get('Royal', 0)}
        except Exception:
            return {'total': 0, 'Basic': 0, 'Glossy': 0, 'Royal': 0}

    def format_copies(self, total, basic, glossy, royal):
        """Format copies into the 'N [Basic A, Glossy B, Royal C]' string."""
        return f"{total} [Basic {basic}, Glossy {glossy}, Royal {royal}]"

    def show_update_screen(self):
        """Open the Update Collection modal for manual edits."""
        if hasattr(self, 'update_window') and self.update_window and self.update_window.winfo_exists():
            return
        self.update_btn['state'] = 'disabled'
        update_window = tk.Toplevel(self.root)
        self.update_window = update_window
        update_window.title("Update Collection")
        update_window.geometry("400x250")
        try:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'favicon.ico')
            if os.path.exists(icon_path):
                update_window.iconbitmap(icon_path)
        except Exception:
            pass
        vcmd = (self.root.register(self.validate_number), '%P')
        self.update_name = tk.StringVar()
        self.update_number = tk.StringVar(value='1')
        self.update_dustable = tk.StringVar(value='1')
        self.update_finish = tk.StringVar(value='Basic')
        
        def on_name_change(*args):
            self.update_number.set('1'); self.update_dustable.set('1'); self.update_finish.set('Basic')
        self.update_name.trace('w', on_name_change)

        def on_number_change(*args):
            count_val = self.update_number.get()
            if count_val:
                dustable_entry['state'] = 'normal'
                max_dust = int(count_val)
                current_dust = int(self.update_dustable.get() or 0)
                if current_dust > max_dust:
                    self.update_dustable.set(str(max_dust))
                else:
                    self.update_dustable.set(count_val)
            else:
                dustable_entry['state'] = 'disabled'; self.update_dustable.set('')

        def on_dustable_change(*args):
            count_val = self.update_number.get()
            if count_val:
                max_dust = int(count_val)
                current_dust = int(self.update_dustable.get() or 0)
                if current_dust > max_dust:
                    self.update_dustable.set(str(max_dust))
        self.update_number.trace('w', on_number_change)
        self.update_dustable.trace('w', on_dustable_change)

        def on_close():
            self.update_window = None
            self.update_btn['state'] = 'normal'
            update_window.destroy()

        update_window.protocol("WM_DELETE_WINDOW", on_close)
        ttk.Label(update_window, text="Card Name:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        name_entry = ttk.Entry(update_window, textvariable=self.update_name); name_entry.grid(row=0, column=1, padx=10, pady=10, sticky='ew')
        ttk.Label(update_window, text="Count:").grid(row=1, column=0, padx=10, pady=10, sticky='w')
        number_entry = ttk.Entry(update_window, textvariable=self.update_number, validate='key', validatecommand=vcmd); number_entry.grid(row=1, column=1, padx=10, pady=10, sticky='ew')
        ttk.Label(update_window, text="Finish:").grid(row=2, column=0, padx=10, pady=10, sticky='w')
        finish_combo = ttk.Combobox(update_window, textvariable=self.update_finish, values=['Basic','Glossy','Royal'], state='readonly'); finish_combo.grid(row=2, column=1, padx=10, pady=10, sticky='ew')
        ttk.Label(update_window, text="Dustable:").grid(row=3, column=0, padx=10, pady=10, sticky='w')
        dustable_entry = ttk.Entry(update_window, textvariable=self.update_dustable, validate='key', validatecommand=vcmd); dustable_entry.grid(row=3, column=1, padx=10, pady=10, sticky='ew')
        buttons_frame = ttk.Frame(update_window); buttons_frame.grid(row=4, column=0, columnspan=2, pady=5)
        add_btn = ttk.Button(buttons_frame, text="ADD", command=self.add_card); add_btn.pack(side='left', padx=10)
        delete_btn = ttk.Button(buttons_frame, text="DELETE", command=self.delete_card); delete_btn.pack(side='right', padx=10)
        update_window.columnconfigure(1, weight=1)

    def add_card(self):
        """Add or update a card entry in the current CSV file."""
        try:
            name = self.update_name.get().strip(); count = int(self.update_number.get() or 0); finish = self.update_finish.get(); dustable = int(self.update_dustable.get() or 0)
            if not name or count <= 0:
                self.load_log("Error: Please enter a card name and positive count.", "error"); return
            from card_name_matcher import get_canonical_name_and_legacy_status
            canonical, legacy = get_canonical_name_and_legacy_status(name)
            if not canonical:
                self.load_log("Error: Card not found.", "error"); return
            df = pd.read_csv(self.current_csv, dtype=str, keep_default_na=False, na_values=[])
            mask = df['Name'] == canonical
            if mask.any():
                idx = df[mask].index[0]
                copies_dict = self.parse_copies(df.at[idx, 'Copies'])
                copies_dict['total'] += count
                copies_dict[finish] += count
                df.at[idx, 'Copies'] = self.format_copies(copies_dict['total'], copies_dict['Basic'], copies_dict['Glossy'], copies_dict['Royal'])
                df.at[idx, 'Dustable'] = str(int(df.at[idx, 'Dustable'] or 0) + dustable)
            else:
                card_info = get_card_info(canonical)
                if "error" in card_info:
                    self.load_log(f"API Error for {canonical}: {card_info['error']}", "error"); return
                if not card_info:
                    self.load_log(f"Error: Failed to get card data for {canonical}.", "error"); return
                card_rarity = get_card_rarity_from_info(card_info)
                frame_type = card_info.get("frameType", "")
                card_type = "Trap" if frame_type == "trap" else "Spell" if frame_type == "spell" else "Monster"
                card_stats = ""
                if card_info.get("type") not in ["Trap Card", "Spell Card"]:
                    level = card_info.get("level", "")
                    attribute = card_info.get("attribute", "")
                    race = card_info.get("race", "")
                    atk = card_info.get("atk", "")
                    def_val = card_info.get("def")
                    def_str = "-" if def_val is None else str(def_val)
                    if frame_type == "xyz":
                        card_stats = f" Rank {level}, "
                    elif frame_type == "link":
                        linkval = card_info.get("linkval", "")
                        arrows = card_info.get("linkmarkers", [])
                        card_stats = f" Link {linkval}, "
                        if arrows:
                            arrow_symbols = [ARROW_MAP.get(dir, "?") for dir in arrows]
                            card_stats += f"{''.join(arrow_symbols)}, "
                    else:
                        card_stats = f" Level {level}, "
                    card_stats += f"{attribute}, {race}, {atk}/{def_str}"
                new_row = {
                    'Rarity': get_card_rarity_from_info(card_info), 'Name': canonical,
                    'Legacy': 'Yes' if legacy else 'No', 'Copies': self.format_copies(count, count if finish=='Basic' else 0, count if finish=='Glossy' else 0, count if finish=='Royal' else 0),
                    'Dustable': str(dustable), 'Archetype': card_info.get('archetype') or '', 'Card Frame': card_type, 'Card Type': card_info.get('humanReadableCardType', ""),
                    'Card Stats': card_stats, 'Effect': card_info.get('desc', "").replace("\n", " ").replace("\r", " ")
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            df.to_csv(self.current_csv, index=False)
            self.load_log(f"Added {count} {finish} copies of {canonical}.", "info")
        except Exception as e:
            self.load_log(f"Failed to add card: {str(e)}", "error")

    def delete_card(self):
        """Remove copies or an entry from the current CSV file."""
        try:
            name = self.update_name.get().strip(); count = int(self.update_number.get() or 0); finish = self.update_finish.get(); dustable = int(self.update_dustable.get() or 0)
            if not name or count <= 0:
                self.load_log("Error: Please enter a card name and positive count.", "error"); return
            from card_name_matcher import get_canonical_name_and_legacy_status
            canonical, _ = get_canonical_name_and_legacy_status(name)
            if not canonical:
                self.load_log("Error: Card not found.", "error"); return
            df = pd.read_csv(self.current_csv, dtype=str, keep_default_na=False, na_values=[])
            mask = df['Name'] == canonical
            if not mask.any():
                self.load_log("Error: Card not in collection.", "error"); return
            idx = df[mask].index[0]
            copies_dict = self.parse_copies(df.at[idx, 'Copies'])
            current_finish = copies_dict[finish]
            if copies_dict['total'] < count or current_finish < count:
                self.load_log("Error: Not enough copies to remove.", "error"); return
            copies_dict['total'] -= count; copies_dict[finish] -= count
            new_dustable = max(0, int(df.at[idx, 'Dustable'] or 0) - dustable)
            if copies_dict['total'] == 0:
                df = df.drop(idx); self.load_log(f"Removed all copies of {canonical}, entry deleted.", "info")
            else:
                df.at[idx, 'Copies'] = self.format_copies(copies_dict['total'], copies_dict['Basic'], copies_dict['Glossy'], copies_dict['Royal'])
                df.at[idx, 'Dustable'] = str(new_dustable)
                self.load_log(f"Removed {count} {finish} copies of {canonical}.", "info")
            df.to_csv(self.current_csv, index=False)
        except Exception as e:
            self.load_log(f"Failed to delete card: {str(e)}", "error")

    def update_status(self, message: str):
        """Log a status message to execution terminal."""
        self.log(f"Status: {message}", "info")

    def log_debug_mode(self):
        """Toggle debug mode logging."""
        self.log("Debug mode ENABLED" if self.debug_mode.get() else "Debug mode DISABLED", "info")

    def log_print_summary(self):
        """Toggle summary printing logging."""
        self.log("Print summary ENABLED" if self.print_summary.get() else "Print summary DISABLED", "info")

    def save_terminal_content(self, terminal, key):
        """Persist terminal contents (text and tags) to restore later."""
        if not terminal:
            return
        terminal.configure(state='normal')
        content = terminal.get('1.0', 'end-1c')
        terminal.configure(state='disabled')
        tags = []
        for tag_name in terminal.tag_names():
            ranges = terminal.tag_ranges(tag_name)
            tags.append((tag_name, list(ranges)))
        self.terminal_contents[key] = (content, tags)

    def restore_terminal_content(self, terminal, key):
        """Restore previously saved terminal contents and tags."""
        if key not in self.terminal_contents:
            return
        content, tags = self.terminal_contents[key]
        terminal.configure(state='normal')
        terminal.delete('1.0', 'end')
        terminal.insert('1.0', content)
        self.root.update_idletasks()
        for tag_name, ranges in tags:
            for i in range(0, len(ranges), 2):
                terminal.tag_add(tag_name, ranges[i], ranges[i+1])
        terminal.see('end')
        terminal.configure(state='disabled')

    def _back_from_create(self):
        self.save_terminal_content(self.execution_terminal, 'execution')
        self.show_main_menu()

    def _back_from_load(self):
        self.save_terminal_content(self.load_terminal, 'load')
        self.show_main_menu()

    def start_collection_scan(self):
        """Spawn the collection scanner subprocess and stream output to terminal."""
        try:
            save_path = self.save_entry.get().strip()
            if not save_path:
                messagebox.showerror("Error", "Please specify a save location")
                self.log("Error: No save location specified", "error")
                return
            if os.path.isfile(save_path):
                save_path = os.path.dirname(save_path)
                self.log(f"Using parent directory: {save_path}", "debug")
            try:
                os.makedirs(save_path, exist_ok=True)
                self.log(f"Using save directory: {save_path}", "info")
            except Exception as e:
                messagebox.showerror("Error", f"Could not create directory: {str(e)}")
                self.log(f"Could not create directory: {str(e)}", "error")
                return
            self.scan_in_progress = True
            self.start_btn['state'] = 'disabled'
            self.stop_btn['state'] = 'normal'
            self.log("Initializing collection scan...", "info")
            cmd = [sys.executable, '-u', 'main-new-better.py']
            if self.debug_mode.get():
                cmd.append('--debug')
            if not self.print_summary.get():
                cmd.append('--no-summary')
            cmd.extend(['--output-dir', save_path, '--game-res', self.game_resolution.get()])
            self.process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, cwd=os.getcwd())
            threading.Thread(target=self.read_output, daemon=True).start()
        except Exception as e:
            self.update_status(f"Error: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.scan_in_progress = False
            self.start_btn['state'] = 'normal'
            self.stop_btn['state'] = 'disabled'
        self.root.update_idletasks()

    def read_output(self):
        """Read subprocess stdout and append lines to execution terminal with ANSI tags."""
        from datetime import datetime
        for line in iter(self.process.stdout.readline, ''):
            clean_line, ranges = parse_ansi(line)
            timestamp = datetime.now().strftime("%H:%M:%S")
            ts_prefix = f"[{timestamp}] "
            full_line = ts_prefix + clean_line
            self.execution_terminal.configure(state='normal')
            start_idx = self.execution_terminal.index('end-1c')
            self.execution_terminal.insert('end', full_line)
            ts_end = len(ts_prefix)
            self.execution_terminal.tag_add('timestamp', f"{start_idx}+0c", f"{start_idx}+{ts_end}c")
            for s, e, tag in ranges:
                adj_s = s + ts_end; adj_e = e + ts_end
                tag_start = f"{start_idx}+{adj_s}c"; tag_end = f"{start_idx}+{adj_e}c"
                self.execution_terminal.tag_add(tag, tag_start, tag_end)
            self.execution_terminal.see('end')
            self.execution_terminal.configure(state='disabled')
            self.root.update_idletasks()
        self.process.stdout.close(); self.process.wait()
        self.scan_in_progress = False
        self.start_btn['state'] = 'normal'; self.stop_btn['state'] = 'disabled'
        self.log("Execution completed.")
        
    def stop_collection_scan(self):
        """Terminate running scanner process."""
        if self.scan_in_progress and self.process:
            self.process.terminate(); self.log("Scan stopped by user.", "warning")
        self.scan_in_progress = False
        self.start_btn['state'] = 'normal'; self.stop_btn['state'] = 'disabled'

def main():
    root = tk.Tk(); app = MasterDuelExporterApp(root); root.mainloop()

if __name__ == "__main__":
    main()
